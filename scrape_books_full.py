#!/usr/bin/env python3
# scrape_books_full.py
import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from urllib.parse import urljoin
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from datetime import datetime

# ---------- Scrape ----------
titles = []
prices = []
ratings = []
availability_list = []
book_urls = []
categories = []

page_num = 1
base_first_page = "https://books.toscrape.com/index.html"

print("Starting scrape...")

while True:
    # page 1 uses index.html; later pages follow catalogue/page-N.html
    if page_num == 1:
        url = base_first_page
    else:
        url = f"https://books.toscrape.com/catalogue/page-{page_num}.html"

    print(f"Scraping page {page_num}: {url}")

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Failed to scrape page {page_num}: {e}")
        break

    soup = BeautifulSoup(response.text, "html.parser")
    books = soup.find_all("article", class_="product_pod")

    if not books:
        print(f"No books found on page {page_num}. Stopping.")
        break

    for book in books:
        # Title
        title = book.h3.a.get("title", "Unknown Title")

        # Price (raw)
        price_raw = book.select_one('p.price_color').text.strip()
        # keep the raw for a moment; we'll clean later
        price_clean = price_raw.replace("Â", "").replace("£", "").strip()

        # Availability
        availability = book.select_one('p.instock.availability').text.strip()

        # Rating (word)
        rating_tag = book.select_one('p.star-rating')
        rating_classes = rating_tag.get("class", []) if rating_tag else []
        star_rating_word = next((c for c in rating_classes if c != "star-rating"), "No Rating")

        # URL for the book (resolve relative paths)
        partial_link = book.h3.a["href"]
        book_url = urljoin(url, partial_link)

        # Get category from detail page (safely)
        category = "Unknown"
        try:
            book_response = requests.get(book_url, timeout=10)
            book_response.raise_for_status()
            book_soup = BeautifulSoup(book_response.text, "html.parser")
            breadcrumb = book_soup.select("ul.breadcrumb li a")
            if len(breadcrumb) > 2:
                category = breadcrumb[2].text.strip()
        except Exception as e:
            # don't fail the whole run for a single book
            print(f"Warning: couldn't fetch details for '{title}': {e}")

        # Append
        titles.append(title)
        prices.append(price_clean)
        ratings.append(star_rating_word)
        availability_list.append(availability)
        book_urls.append(book_url)
        categories.append(category)

        # polite short pause for the book detail request
        time.sleep(0.2)

    page_num += 1
    time.sleep(0.8)  # polite delay between pages

print("Scrape finished. Preparing dataframe...")

# ---------- DataFrame & Cleaning ----------
df = pd.DataFrame({
    "Title": titles,
    "Price_raw": prices,
    "Rating_word": ratings,
    "Availability": availability_list,
    "URL": book_urls,
    "Category": categories
})

# Clean price -> numeric float
df["Price"] = df["Price_raw"].str.replace("Â", "", regex=False).str.replace("£", "", regex=False)
# If any prices have commas or weird chars, remove them and convert
df["Price"] = df["Price"].str.replace(",", "", regex=False)
df["Price"] = pd.to_numeric(df["Price"], errors="coerce")

# Convert rating words to numbers
rating_map = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}
df["Rating"] = df["Rating_word"].map(rating_map).fillna(0).astype(int)

# Tidy availability (remove newlines/extra spaces)
df["Availability"] = df["Availability"].str.replace("\n", " ", regex=False).str.strip()

# Reorder & drop helper columns
df = df[["Title", "Price", "Rating", "Availability", "Category", "URL", "Price_raw", "Rating_word"]]

# Save a clean CSV (raw data for clients)
df.to_csv("books_clean.csv", index=False, encoding="utf-8")
print("Saved books_clean.csv")

# ---------- Styled Excel Export ----------
excel_path = "Books_Styled.xlsx"
# Save current df to excel (first create Excel)
df_to_save = df.copy()
# Optionally drop helper cols from the view. We'll show Price_raw and Rating_word in hidden cols maybe; for now export main cols only
df_to_save_display = df_to_save[["Title", "Price", "Rating", "Availability", "Category", "URL"]]
df_to_save_display.to_excel(excel_path, index=False)

# Load workbook
wb = load_workbook(excel_path)

# Create a Cover sheet at index 0
cover = wb.create_sheet(index=0, title="Cover")
cover_title = "Books Scraper — Project Deliverable"
cover["A1"] = cover_title
cover["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
cover["A4"] = "Description:"
cover["A5"] = "This dataset was generated by a Python web scraper (requests + BeautifulSoup)."
cover["A6"] = "It includes book Title, Price (GBP), Rating (1-5), Availability, Category and direct URL to the product page."
cover["A8"] = "Contact:"
cover["A9"] = "Name: [Your Name Here]"
cover["A10"] = "Email: [your.email@example.com]"

# Simple formatting for cover
cover.column_dimensions["A"].width = 60
cover["A1"].font = Font(size=18, bold=True)
cover["A1"].alignment = Alignment(horizontal="left", vertical="center")
for cell in ["A2", "A4", "A5", "A6", "A8", "A9", "A10"]:
    cover[cell].alignment = Alignment(wrap_text=True)

# Style the data sheet (it should be the second sheet now)
ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[1]]

# Header styling
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Price formatting and center Ratings
price_col_letter = None
rating_col_letter = None
# find the columns (top row contains headers)
for idx, cell in enumerate(ws[1], start=1):
    header = str(cell.value).strip().lower()
    if header == "price":
        price_col_letter = get_column_letter(idx)
    if header == "rating":
        rating_col_letter = get_column_letter(idx)

# Apply number format to Price and center rating
if price_col_letter:
    for row in range(2, ws.max_row + 1):
        try:
            ws[f"{price_col_letter}{row}"].number_format = '£#,##0.00'
        except Exception:
            pass

if rating_col_letter:
    for row in range(2, ws.max_row + 1):
        ws[f"{rating_col_letter}{row}"].alignment = Alignment(horizontal="center")

# Auto-fit column widths
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except Exception:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Freeze header row
ws.freeze_panes = "A2"

# Move the data sheet to be after Cover (optional)
# Save workbook
wb.save(excel_path)
print(f"Saved styled Excel as {excel_path}")

print("All done — books_clean.csv and Books_Styled.xlsx created.")
